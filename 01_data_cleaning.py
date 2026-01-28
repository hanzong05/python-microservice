"""
Geotechnical Data Extraction and Cleaning
- Downloads Raw_Data.xlsx from Supabase Storage
- Processes all depth layers into Cleaned_Data.xlsx
- Archives existing cleaned files to old_cleaned_data/<timestamp>
- Uploads new cleaned file back to Supabase Storage
"""

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
import io
from supabase_client import get_supabase_client
from datetime import datetime

# -----------------------------
# Download File from Supabase
# -----------------------------


def download_file_from_storage(bucket_name, file_path):
    print(f" Downloading file from Supabase Storage...")
    print(f"   Bucket: {bucket_name}")
    print(f"   File: {file_path}")

    client = get_supabase_client()
    if not client:
        print(" Failed to connect to Supabase")
        return None

    try:
        response = client.storage.from_(bucket_name).download(file_path)
        print(f"✓ Successfully downloaded {len(response)} bytes")
        return response
    except Exception as e:
        print(f" Error downloading file: {e}")
        return None

# -----------------------------
# Process Excel Data
# -----------------------------


def process_geotechnical_data(file_bytes, output_filename='Cleaned_Data.xlsx'):
    print(f"\n Processing geotechnical data...")
    wb = load_workbook(io.BytesIO(file_bytes))

    sheet_names = ['0m-1.5m', '1.5m-3.0m', '3.0m-4.5m', '4.5m-6.0m', '6.0m-7.5m',
                   '7.5m-9.0m', '9.0m-10.5m', '10.5m-12.0m', '12.0m-13.5m', '13.5m-15.0m']

    all_data = []

    for sheet_name in sheet_names:
        print(f"  Processing layer: {sheet_name}...")
        sheet = wb[sheet_name]

        data = []
        headers = None
        for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if i == 2:
                headers = row
            elif i > 2:
                data.append(row)

        df = pd.DataFrame(data, columns=headers)
        df['Depth_Layer'] = sheet_name
        # Clean latitude/longitude
        df['Latitude'] = df['Latitude'].astype(str).str.replace('°', '')
        df['Longitude'] = df['Longitude'].astype(str).str.replace('°', '')
        all_data.append(df)

    combined_df = pd.concat(all_data, ignore_index=True)
    print(
        f"\n✓ Extracted {len(combined_df)} records from {len(sheet_names)} depth layers")

    # -----------------------------
    # Create Cleaned Workbook
    # -----------------------------
    new_wb = Workbook()
    sheet = new_wb.active
    sheet.title = "Cleaned_Data"

    header_fill = PatternFill(start_color='366092',
                              end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')

    for col_idx, header in enumerate(combined_df.columns, 1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row_idx, row_data in enumerate(combined_df.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    sheet.freeze_panes = 'A2'

    # Summary sheet
    summary_sheet = new_wb.create_sheet('Summary')
    summary_sheet['A1'] = 'Geotechnical Data Extraction Summary'
    summary_sheet['A1'].font = Font(bold=True, size=14)
    summary_sheet['A3'] = 'Total Records:'
    summary_sheet['B3'] = len(combined_df)
    summary_sheet['A4'] = 'Depth Layers:'
    summary_sheet['B4'] = len(sheet_names)
    summary_sheet['A5'] = 'Municipalities:'
    summary_sheet['B5'] = combined_df['Municipality'].nunique(
    ) if 'Municipality' in combined_df.columns else 'N/A'
    summary_sheet['A6'] = 'Total Boreholes:'
    summary_sheet['B6'] = combined_df['Borehole ID'].nunique(
    ) if 'Borehole ID' in combined_df.columns else 'N/A'
    summary_sheet['A8'] = 'Depth Layers Included:'
    for i, layer in enumerate(sheet_names, 9):
        summary_sheet[f'A{i}'] = layer

    summary_sheet.column_dimensions['A'].width = 30
    summary_sheet.column_dimensions['B'].width = 20

    new_wb.save(output_filename)
    print(f"\n✓ Output file saved: {output_filename}")
    return combined_df

# -----------------------------
# Upload & Archive Existing File
# -----------------------------


def upload_to_supabase_storage(local_file_path, bucket_name, storage_path):
    print(f"\n Uploading to Supabase Storage...")
    client = get_supabase_client()
    if not client:
        return False

    try:
        # Check if file already exists
        try:
            existing_file = client.storage.from_(
                bucket_name).download(storage_path)
            if existing_file:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                archive_path = f"old_cleaned_data/old_cleaned_data_{timestamp}.xlsx"
                client.storage.from_(bucket_name).move(
                    storage_path, archive_path)
                print(f"✓ Existing file archived as {archive_path}")
        except Exception as e:
            # File may not exist
            pass

        # Upload new file
        with open(local_file_path, 'rb') as f:
            file_data = f.read()
        client.storage.from_(bucket_name).upload(
            storage_path,
            file_data,
            file_options={
                "content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        )
        print(f"✓ File uploaded successfully to {storage_path}")
        return True

    except Exception as e:
        print(f" Error uploading file: {e}")
        return False

# -----------------------------
# Main Execution
# -----------------------------


def main():
    print("="*70)
    print("GEOTECHNICAL DATA EXTRACTION & CLEANING")
    print("="*70)

    BUCKET_NAME = 'geotechnical-data'
    INPUT_FILE_PATH = 'raw/Raw_Data.xlsx'           # Source in storage
    OUTPUT_FILE_NAME = 'Cleaned_Data.xlsx'
    OUTPUT_STORAGE_PATH = 'cleaned/Cleaned_Data.xlsx'  # Destination in storage

    # Download
    file_bytes = download_file_from_storage(BUCKET_NAME, INPUT_FILE_PATH)
    if not file_bytes:
        print("\n Failed to download file")
        return

    # Process
    combined_df = process_geotechnical_data(file_bytes, OUTPUT_FILE_NAME)
    print(f"\n✓ Total Records Processed: {len(combined_df)}")

    # Upload
    upload_to_supabase_storage(
        OUTPUT_FILE_NAME, BUCKET_NAME, OUTPUT_STORAGE_PATH)
    print("\ PROCESSING COMPLETED SUCCESSFULLY!")
    print("="*70)


if __name__ == "__main__":
    main()
